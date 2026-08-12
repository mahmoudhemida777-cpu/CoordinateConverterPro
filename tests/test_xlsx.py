import openpyxl
import pytest

from core.parsers.xlsx_parser import sniff_columns, parse_xlsx, ColumnMapping
from core.exporters.xlsx_exporter import export_xlsx
from core.models import PointResult


@pytest.fixture()
def sample_xlsx(tmp_path):
    path = tmp_path / "points.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Point Name", "Easting", "Northing", "Elevation"])
    ws.append(["P1", 46.84550218, 24.77373641, 650.5])
    ws.append(["P2", 46.85, 24.78, 651.0])
    wb.save(path)
    return str(path)


def test_sniff_columns(sample_xlsx):
    assert sniff_columns(sample_xlsx) == ["Point Name", "Easting", "Northing", "Elevation"]


def test_parse_xlsx_points(sample_xlsx):
    mapping = ColumnMapping(name_col="Point Name", x_col="Easting", y_col="Northing", z_col="Elevation")
    points = parse_xlsx(sample_xlsx, mapping)
    assert len(points) == 2
    assert points[0].src_x == pytest.approx(46.84550218)


def test_export_xlsx_creates_points_and_project_info_sheets(tmp_path):
    out = tmp_path / "export.xlsx"
    pts = [PointResult("P1", 46.8455, 24.7737, 650.5, 500000.1, 2740123.4, 650.5, "SUCCESS")]
    export_xlsx(pts, str(out), "EPSG:4326", "EPSG:20438", {"datum": "Ain el Abd 1970"})
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Points", "Project Info"]


def test_export_xlsx_points_sheet_headers(tmp_path):
    out = tmp_path / "export.xlsx"
    pts = [PointResult("P1", 46.8455, 24.7737, 650.5, 500000.1, 2740123.4, 650.5, "SUCCESS")]
    export_xlsx(pts, str(out), "EPSG:4326", "EPSG:20438")
    wb = openpyxl.load_workbook(out)
    ws = wb["Points"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header == ["No.", "Point Name", "Source X", "Source Y", "Source Z",
                       "Target X", "Target Y", "Target Z", "Target X,Y", "Status"]


def test_export_xlsx_project_info_contains_required_fields(tmp_path):
    out = tmp_path / "export.xlsx"
    pts = [PointResult("P1", 1.0, 1.0, None, 1.0, 1.0, None, "SUCCESS")]
    export_xlsx(pts, str(out), "EPSG:4326", "EPSG:20438")
    wb = openpyxl.load_workbook(out)
    ws = wb["Project Info"]
    labels = [row[0].value for row in ws.iter_rows(min_row=1, max_col=1)]
    for required in ["Source CRS", "Target CRS", "Datum", "Projection", "Ellipsoid",
                      "Units", "Number of Points", "Date", "Software Version"]:
        assert required in labels
