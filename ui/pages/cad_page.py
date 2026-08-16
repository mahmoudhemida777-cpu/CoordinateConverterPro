from __future__ import annotations
import csv
from pathlib import Path
import openpyxl
from PySide6.QtWidgets import QWidget,QVBoxLayout,QHBoxLayout,QLabel,QPushButton,QFileDialog,QTableWidget,QTableWidgetItem,QProgressBar,QMessageBox,QGroupBox,QCheckBox,QComboBox,QRadioButton,QFormLayout,QScrollArea,QGridLayout,QHeaderView,QSizePolicy
from core.crs.engine import CRSEngine
from core.models import PointResult
from core.parsers import csv_parser,xlsx_parser,kml_parser,txt_parser
from core.exporters.dxf_exporter import export_dxf,LabelMode
from core.point_ordering import order_points
from ui.i18n import tr
from ui.widgets.crs_picker import CRSPicker
from ui.widgets.workspace_bar import WorkspaceFileBar
from ui.pages.settings_page import current_precision

COORDINATE_FILTER="Coordinate files (*.kmz *.kml *.csv *.xlsx *.txt);;KMZ/KML (*.kmz *.kml);;CSV (*.csv);;Excel (*.xlsx);;Survey TXT (*.txt);;All files (*.*)"

class CadPage(QWidget):
    def __init__(self)->None:
        super().__init__(); self.engine=CRSEngine(); self.source_points=[]; self.result_points=[]; self.current_file=None; self.batch_converted=False; self.workspace_folder=None; self._detected_columns=[]; self._axis_swapped=False
        outer=QVBoxLayout(self); outer.setContentsMargins(0,0,0,0); outer.setSpacing(0)
        scroll=QScrollArea(); scroll.setWidgetResizable(True); scroll.setFrameShape(QScrollArea.Shape.NoFrame); outer.addWidget(scroll)
        content=QWidget(); root=QVBoxLayout(content); root.setContentsMargins(26,18,26,26); root.setSpacing(12); scroll.setWidget(content)
        title=QLabel("AutoCAD / Civil 3D Export"); title.setStyleSheet("font-size:20px;font-weight:bold;color:#1F3864;margin-bottom:2px;"); root.addWidget(title)
        subtitle=QLabel("Load ordinary survey points and export them directly to CAD, or convert between CRS before export."); subtitle.setWordWrap(True); subtitle.setStyleSheet("color:#555;"); root.addWidget(subtitle)
        self.workspace_bar=WorkspaceFileBar(); self.workspace_bar.file_selected.connect(self._load_path); root.addWidget(self.workspace_bar)
        row=QHBoxLayout(); row.setSpacing(10); choose=QPushButton("Choose Coordinate File"); choose.setMinimumHeight(34); choose.clicked.connect(self._choose_file); row.addWidget(choose,0); self.file_label=QLabel("No file selected"); self.file_label.setStyleSheet("color:#777;"); self.file_label.setWordWrap(True); row.addWidget(self.file_label,1); root.addLayout(row)
        self.direct_mode=QCheckBox("DIRECT CAD EXPORT — use the loaded coordinates exactly as they are (NO CRS conversion)"); self.direct_mode.setChecked(False); self.direct_mode.stateChanged.connect(self._mode_changed); root.addWidget(self.direct_mode)

        parsing=QGroupBox("FILE PARSING OPTIONS"); parsing.setStyleSheet("QGroupBox{font-weight:bold;color:#1F3864;border:1px solid #C9D2DE;border-radius:8px;margin-top:8px;padding:12px;} QGroupBox::title{subcontrol-origin:margin;left:12px;padding:0 5px;background:white;} QLabel{font-weight:normal;color:#333;}"); parsing_layout=QVBoxLayout(parsing); parsing_layout.setSpacing(10)
        top_grid=QGridLayout(); top_grid.setHorizontalSpacing(12); top_grid.setVerticalSpacing(8)
        self.parsing_engine=QComboBox(); self.parsing_engine.addItems(["Smart (Recommended)","Manual / Selected Columns"]); self.parsing_engine.setMinimumHeight(30)
        self.detected_format=QComboBox(); self.detected_format.setEnabled(False); self.detected_format.setMinimumHeight(30)
        top_grid.addWidget(QLabel("Parsing Engine"),0,0); top_grid.addWidget(self.parsing_engine,0,1); top_grid.addWidget(QLabel("Detected Format"),0,2); top_grid.addWidget(self.detected_format,0,3)
        self.x_column=QComboBox(); self.y_column=QComboBox(); self.z_column=QComboBox(); self.name_column=QComboBox()
        for combo in (self.x_column,self.y_column,self.z_column,self.name_column): combo.setMinimumHeight(30); combo.setSizePolicy(QSizePolicy.Policy.Expanding,QSizePolicy.Policy.Fixed)
        top_grid.addWidget(QLabel("Easting / X"),1,0); top_grid.addWidget(self.x_column,1,1); top_grid.addWidget(QLabel("Northing / Y"),1,2); top_grid.addWidget(self.y_column,1,3)
        top_grid.addWidget(QLabel("Elevation / Z"),2,0); top_grid.addWidget(self.z_column,2,1); top_grid.addWidget(QLabel("Point Name / Code"),2,2); top_grid.addWidget(self.name_column,2,3)
        top_grid.setColumnStretch(1,1); top_grid.setColumnStretch(3,1); parsing_layout.addLayout(top_grid)

        axis_box=QGroupBox("AXIS ORDER — IMPORTANT"); axis_box.setStyleSheet("QGroupBox{font-weight:bold;color:#8A5A00;border:1px solid #D8C58A;border-radius:7px;padding:8px;margin-top:2px;} QRadioButton{font-weight:normal;color:#333;padding:3px;}"); af=QHBoxLayout(axis_box); af.setContentsMargins(10,6,10,6); af.setSpacing(20)
        self.axis_xy=QRadioButton("Easting (X) → Northing (Y) — CAD standard"); self.axis_yx=QRadioButton("Northing (Y) → Easting (X) — SWAP"); self.axis_xy.setChecked(True); af.addWidget(self.axis_xy); af.addWidget(self.axis_yx); af.addStretch(); parsing_layout.addWidget(axis_box)

        ordering_box=QGroupBox("GRID / ZIGZAG POINT NUMBERING"); ordering_box.setStyleSheet("QGroupBox{font-weight:bold;color:#1F3864;border:1px solid #C9D2DE;border-radius:7px;padding:8px;margin-top:2px;} QLabel{font-weight:normal;color:#333;} QComboBox{min-height:30px;}"); of=QHBoxLayout(ordering_box); of.setContentsMargins(10,6,10,6); of.setSpacing(12)
        of.addWidget(QLabel("Numbering / Ordering")); self.ordering_mode=QComboBox(); self.ordering_mode.addItem("Grid Zigzag — Start West (W → E)","GRID_ZIGZAG_WEST"); self.ordering_mode.addItem("Grid Zigzag — Start East (E → W)","GRID_ZIGZAG_EAST"); self.ordering_mode.addItem("Keep Source Order","SOURCE"); of.addWidget(self.ordering_mode,1); of.addWidget(QLabel("First row starts from the selected side; each next row reverses direction.")); parsing_layout.addWidget(ordering_box)

        action_box=QHBoxLayout(); action_box.setSpacing(8); preview=QPushButton("Preview"); apply_all=QPushButton("Apply to All"); reset=QPushButton("Reset");
        for button in (preview,apply_all,reset): button.setMinimumHeight(32)
        preview.clicked.connect(self._apply_parsing_options); apply_all.clicked.connect(self._apply_parsing_options); reset.clicked.connect(self._reset_parsing_options); action_box.addWidget(preview); action_box.addWidget(apply_all); action_box.addWidget(reset); action_box.addStretch(); parsing_layout.addLayout(action_box); root.addWidget(parsing)

        crs_group=QGroupBox("COORDINATE REFERENCE SYSTEM"); crs_group.setStyleSheet("QGroupBox{font-weight:bold;color:#1F3864;border:1px solid #C9D2DE;border-radius:8px;margin-top:8px;padding:10px;} QGroupBox::title{subcontrol-origin:margin;left:12px;padding:0 5px;background:white;}"); crs_row=QHBoxLayout(crs_group); crs_row.setSpacing(12); self.source_picker=CRSPicker(self.engine,tr("SOURCE CRS")); self.target_picker=CRSPicker(self.engine,tr("TARGET CRS")); crs_row.addWidget(self.source_picker,1); crs_row.addWidget(self.target_picker,1); root.addWidget(crs_group)
        convert=QPushButton("CONVERT FOR CAD / CIVIL 3D"); convert.setMinimumHeight(40); convert.setStyleSheet("background-color:#C9A227;color:white;font-weight:bold;padding:9px 24px;border-radius:5px;"); convert.clicked.connect(self._convert); root.addWidget(convert)
        self.progress=QProgressBar(); self.progress.setMinimumHeight(20); root.addWidget(self.progress)
        summary=QGroupBox(); sl=QHBoxLayout(summary); sl.setContentsMargins(10,7,10,7); self.total=QLabel("Points: 0"); self.success=QLabel("Success: 0"); self.failed=QLabel("Failed: 0"); [sl.addWidget(w) for w in (self.total,self.success,self.failed)]; sl.addStretch(); root.addWidget(summary)
        self.table=QTableWidget(0,6); self.table.setHorizontalHeaderLabels(["Point","Easting / X","Northing / Y","Elevation","Status","Message"]); self.table.setMinimumHeight(240); self.table.horizontalHeader().setSectionResizeMode(0,QHeaderView.ResizeMode.ResizeToContents); self.table.horizontalHeader().setSectionResizeMode(1,QHeaderView.ResizeMode.ResizeToContents); self.table.horizontalHeader().setSectionResizeMode(2,QHeaderView.ResizeMode.ResizeToContents); self.table.horizontalHeader().setSectionResizeMode(3,QHeaderView.ResizeMode.ResizeToContents); self.table.horizontalHeader().setSectionResizeMode(4,QHeaderView.ResizeMode.ResizeToContents); self.table.horizontalHeader().setSectionResizeMode(5,QHeaderView.ResizeMode.Stretch); self.table.setAlternatingRowColors(True); root.addWidget(self.table)
        export_row=QHBoxLayout(); export_row.setSpacing(10); dxf=QPushButton("EXPORT DXF — AutoCAD / Civil 3D"); civil=QPushButton("EXPORT CSV — Civil 3D Points"); dxf.setMinimumHeight(38); civil.setMinimumHeight(38); dxf.clicked.connect(self._export_dxf); civil.clicked.connect(self._export_civil3d_csv); export_row.addWidget(dxf,1); export_row.addWidget(civil,1); root.addLayout(export_row)

    def set_workspace_folder(self,folder:str): self.workspace_folder=folder; self.workspace_bar.set_folder(folder,self.current_file)
    def load_active_file(self,path:str)->None:
        if path and Path(path).is_file(): self.workspace_folder=str(Path(path).parent); self.workspace_bar.set_folder(self.workspace_folder,path); self._load_path(path)
    def _choose_file(self)->None:
        path,_=QFileDialog.getOpenFileName(self,"Choose Coordinate File",self.workspace_folder or "",COORDINATE_FILTER)
        if path:self._load_path(path)
    def _mode_changed(self)->None:
        direct=self.direct_mode.isChecked(); self.source_picker.setEnabled(not direct); self.target_picker.setEnabled(not direct)
        if direct and self.source_points:
            self.result_points=[PointResult(p.name,p.src_x,p.src_y,p.src_z,p.src_x,p.src_y,p.src_z,status="SUCCESS",message="DIRECT — no CRS conversion") for p in self.source_points]
            self.progress.setMaximum(len(self.result_points)); self.progress.setValue(len(self.result_points)); self._populate()
        elif not direct and self.result_points and all("DIRECT" in (p.message or "") for p in self.result_points):
            self.result_points=[]; self._populate()

    @staticmethod
    def _auto_xlsx_mapping(columns:list[str])->xlsx_parser.ColumnMapping:
        normalized={str(c).strip().lower():c for c in columns}
        def pick(names):
            for name in names:
                if name in normalized:return normalized[name]
            for key,value in normalized.items():
                if any(name in key for name in names):return value
            return None
        x=pick(("easting","east","longitude","lon","x")); y=pick(("northing","north","latitude","lat","y")); z=pick(("elevation","elev","height","z")); name=pick(("point number","point_number","point","name","id"))
        if x is None or y is None:
            if len(columns)>=3:x,y=columns[0],columns[1]
            else:raise ValueError("Could not automatically identify X/Easting and Y/Northing columns in the Excel file.")
        return xlsx_parser.ColumnMapping(name_col=name,x_col=x,y_col=y,z_col=z)

    def _populate_parsing_options(self,path:str,suffix:str)->None:
        self._detected_columns=[]; self.detected_format.clear(); self.x_column.clear(); self.y_column.clear(); self.z_column.clear(); self.name_column.clear()
        self.detected_format.addItem({".csv":"CSV (Comma delimited)",".xlsx":"Excel Workbook",".txt":"Survey TXT",".kml":"KML",".kmz":"KMZ"}.get(suffix,suffix))
        if suffix==".csv":
            cols=csv_parser.sniff_columns(path); self._detected_columns=list(cols)
        elif suffix==".xlsx":
            cols=xlsx_parser.sniff_columns(path); self._detected_columns=list(cols)
        else:
            return
        for combo in (self.x_column,self.y_column,self.z_column,self.name_column): combo.addItems(self._detected_columns)
        self.z_column.insertItem(0,"<none>"); self.name_column.insertItem(0,"<none>")
        def pick(names):
            norm=[str(c).strip().lower() for c in self._detected_columns]
            for n in names:
                if n in norm:return norm.index(n)
            for i,k in enumerate(norm):
                if any(n in k for n in names):return i
            return -1
        xi=pick(("easting","east","x","longitude","lon")); yi=pick(("northing","north","y","latitude","lat")); zi=pick(("elevation","elev","height","z")); ni=pick(("point number","point_number","pointid","pointcode","code","point","name","id"))
        if xi>=0:self.x_column.setCurrentIndex(xi)
        if yi>=0:self.y_column.setCurrentIndex(yi)
        if zi>=0:self.z_column.setCurrentIndex(zi+1)
        if ni>=0:self.name_column.setCurrentIndex(ni+1)
        self.axis_xy.setChecked(True); self.axis_yx.setChecked(False); self._axis_swapped=False

    def _apply_parsing_options(self)->None:
        if not self.current_file or not self.source_points:return
        path=self.current_file; suffix=Path(path).suffix.casefold()
        try:
            if suffix==".csv" and self.parsing_engine.currentIndex()==1:
                x=self.x_column.currentText(); y=self.y_column.currentText(); z=self.z_column.currentText(); n=self.name_column.currentText(); mapping=csv_parser.ColumnMapping(None if n=="<none>" else n,x,y,None if z=="<none>" else z); points=csv_parser.parse_csv(path,mapping)
            elif suffix==".xlsx" and self.parsing_engine.currentIndex()==1:
                x=self.x_column.currentText(); y=self.y_column.currentText(); z=self.z_column.currentText(); n=self.name_column.currentText(); mapping=xlsx_parser.ColumnMapping(None if n=="<none>" else n,x,y,None if z=="<none>" else z); points=xlsx_parser.parse_xlsx(path,mapping)
            else:
                if suffix==".csv": points=csv_parser.parse_csv_auto(path)
                elif suffix==".xlsx": points=xlsx_parser.parse_xlsx_auto(path)
                elif suffix==".txt": points=txt_parser.parse_txt(path)
                elif suffix==".kml": points=kml_parser.parse_kml_file(path)
                elif suffix==".kmz": points=kml_parser.parse_kmz_file(path)
                else: raise ValueError(f"Unsupported file type: {suffix}")
            if not points: raise ValueError("No coordinate points found after applying parsing options.")
            if self.axis_yx.isChecked():
                points=[PointResult(p.name,p.src_y,p.src_x,p.src_z,p.tgt_y,p.tgt_x,p.tgt_z,status=p.status,message=(p.message or "")+" | AXIS SWAPPED") for p in points]
                self._axis_swapped=True
            else:self._axis_swapped=False
            self.source_points=list(points); self.result_points=[]; self._populate(); self.file_label.setText(f"{Path(path).name} — {len(points)} points loaded — parsing applied")
            if self.direct_mode.isChecked():self._mode_changed()
        except Exception as exc:QMessageBox.critical(self,"Parsing Error",str(exc))

    def _reset_parsing_options(self)->None:
        self.parsing_engine.setCurrentIndex(0); self.axis_xy.setChecked(True); self.axis_yx.setChecked(False); self._axis_swapped=False; self.ordering_mode.setCurrentIndex(0)
        if self.current_file:self._load_path(self.current_file)

    def _load_path(self,path:str)->None:
        file_path=Path(path)
        if file_path.stem.endswith("_converted") and file_path.suffix.casefold()==".xlsx" and self._load_batch_converted_xlsx(path): return
        try:
            suffix=file_path.suffix.casefold()
            if suffix==".kmz":points=kml_parser.parse_kmz_file(path)
            elif suffix==".kml":points=kml_parser.parse_kml_file(path)
            elif suffix==".csv":points=csv_parser.parse_csv_auto(path)
            elif suffix==".xlsx":points=xlsx_parser.parse_xlsx_auto(path)
            elif suffix==".txt":points=txt_parser.parse_txt(path)
            else:raise ValueError(f"Unsupported file type: {suffix}")
            if not points:raise ValueError("No coordinate points were found in the selected file.")
        except Exception as exc:QMessageBox.critical(self,"Import Error",str(exc));return
        self.batch_converted=False; self.source_points=list(points); self.result_points=[]; self.current_file=str(file_path); self.file_label.setText(f"{file_path.name} — {len(points)} points loaded"); self.table.clearContents(); self.table.setRowCount(0); self.total.setText(f"Points: {len(points)}"); self.success.setText("Success: 0"); self.failed.setText("Failed: 0"); self.progress.setMaximum(len(points)); self.progress.setValue(0)
        self._populate_parsing_options(path,suffix)
        if suffix in {".kml",".kmz"}: self.source_picker.set_selected("EPSG:4326","WGS 84 — Geographic 2D (Latitude / Longitude)")
        if self.direct_mode.isChecked(): self._mode_changed()

    def _load_batch_converted_xlsx(self,path:str)->bool:
        try:
            wb=openpyxl.load_workbook(path,read_only=True,data_only=True); ws=wb["Points"] if "Points" in wb.sheetnames else wb.active; rows=list(ws.iter_rows(values_only=True)); info=wb["Project Info"] if "Project Info" in wb.sheetnames else None; target_crs=None
            if info:
                for row in info.iter_rows(values_only=True):
                    if row and str(row[0]).strip()=="Target CRS" and len(row)>1:target_crs=str(row[1]).strip()
            wb.close()
            if not rows:return False
            headers=[str(x).strip() if x is not None else "" for x in rows[0]]; idx={h:i for i,h in enumerate(headers)}
            if not {"Point Name","Target X","Target Y"}.issubset(idx):return False
            pts=[]
            for n,row in enumerate(rows[1:],1):
                try:
                    sx=float(row[idx["Source X"]]) if "Source X" in idx and row[idx["Source X"]] is not None else None; sy=float(row[idx["Source Y"]]) if "Source Y" in idx and row[idx["Source Y"]] is not None else None; sz=float(row[idx["Source Z"]]) if "Source Z" in idx and row[idx["Source Z"]] is not None else None; x=float(row[idx["Target X"]]); y=float(row[idx["Target Y"]]); z=float(row[idx["Target Z"]]) if "Target Z" in idx and row[idx["Target Z"]] is not None else None; name=str(row[idx["Point Name"]]) if row[idx["Point Name"]] is not None else f"PT-{n}"; status=str(row[idx["Status"]]) if "Status" in idx else "SUCCESS"; message=str(row[idx["Message"]]) if "Message" in idx and row[idx["Message"]] else ""; pts.append(PointResult(name,sx,sy,sz,x,y,z,status=status,message=message))
                except (TypeError,ValueError,IndexError):continue
            if not pts:return False
            if sum(p.status=="SUCCESS" for p in pts)==0:QMessageBox.warning(self,"Invalid Batch Result","The batch workbook contains no successful transformed points.");return True
            self.batch_converted=True; self.source_points=pts; self.result_points=pts; self.current_file=path; self.file_label.setText(f"{Path(path).name} — {len(pts)} points — ALREADY CONVERTED"+(f" — {target_crs}" if target_crs else "")); self._set_crs_both(target_crs); self.progress.setMaximum(len(pts)); self.progress.setValue(len(pts)); self._populate(); return True
        except Exception as exc:QMessageBox.critical(self,"Batch Result Error",f"Could not load batch-converted file:\n{exc}");return False
    def _set_crs_both(self,epsg):
        if not epsg:return
        try:info=self.engine.get_crs_details(epsg);name=info.get("name",epsg)
        except Exception:name=epsg
        self.source_picker.set_selected(epsg,name);self.target_picker.set_selected(epsg,name)
    def _convert(self)->None:
        if self.direct_mode.isChecked():self._mode_changed();return
        if self.batch_converted:QMessageBox.information(self,"Already Converted","This file was produced by Batch Converter and is already in the target CRS. Export it directly to DXF or Civil 3D CSV.");return
        if not self.source_points:QMessageBox.warning(self,"No data","Choose a coordinate file first.");return
        src=self.source_picker.selected_epsg();tgt=self.target_picker.selected_epsg()
        if not src or not tgt:QMessageBox.warning(self,"No CRS","Select both Source CRS and Target CRS.");return
        self.result_points=[];self.progress.setMaximum(len(self.source_points))
        for i,point in enumerate(self.source_points,1):self.result_points.append(self.engine.transform_points(src,tgt,[point])[0]);self.progress.setValue(i)
        self._populate()
    def _populate(self):
        pts=self.result_points if self.result_points else self.source_points;precision=current_precision();self.total.setText(f"Points: {len(self.source_points) if self.source_points else len(pts)}");self.success.setText(f"Success: {sum(p.status=='SUCCESS' for p in pts)}");self.failed.setText(f"Failed: {sum(p.status=='FAILED' for p in pts)}");self.table.setRowCount(len(pts))
        def fmt(v):return "" if v is None else (f"{v:.{precision}f}" if isinstance(v,(int,float)) else str(v))
        for i,p in enumerate(pts):
            x=p.tgt_x if p.tgt_x is not None else p.src_x; y=p.tgt_y if p.tgt_y is not None else p.src_y; z=p.tgt_z if p.tgt_z is not None else p.src_z
            for j,value in enumerate([p.name,x,y,z,p.status,p.message]):self.table.setItem(i,j,QTableWidgetItem(fmt(value)))
    def _selected_order_mode(self)->str:
        return str(self.ordering_mode.currentData() or "GRID_ZIGZAG_WEST")
    def _export_dxf(self)->None:
        if self.direct_mode.isChecked() and self.source_points:
            valid=[p for p in self.source_points if p.src_x is not None and p.src_y is not None];export_points=[PointResult(p.name,p.src_x,p.src_y,p.src_z,p.src_x,p.src_y,p.src_z,status="SUCCESS",message="DIRECT — no CRS conversion") for p in valid]
            if not export_points:QMessageBox.warning(self,"Nothing to export","No valid X/Y coordinates found in the selected file.");return
        else:valid=[p for p in self.result_points if p.status=="SUCCESS" and p.tgt_x is not None and p.tgt_y is not None];export_points=valid
        if not export_points:QMessageBox.warning(self,"Nothing to export","There are no validated coordinates to export.");return
        path,_=QFileDialog.getSaveFileName(self,"Export DXF","CAD_Points.dxf","AutoCAD DXF (*.dxf)")
        if not path:return
        try:export_dxf(export_points,path,label_mode=LabelMode.NAME,text_height=1.0,use_target_coords=True,order_mode=self._selected_order_mode());QMessageBox.information(self,"DXF Exported",f"DXF created successfully:\n{path}")
        except Exception as exc:QMessageBox.critical(self,"DXF Export Error",str(exc))
    def _export_civil3d_csv(self)->None:
        if self.direct_mode.isChecked() and self.source_points:
            valid=[p for p in self.source_points if p.src_x is not None and p.src_y is not None];rows=[(p.src_x,p.src_y,p.src_z,p.name) for p in valid]
        else:
            valid=[p for p in self.result_points if p.status=="SUCCESS" and p.tgt_x is not None and p.tgt_y is not None];rows=[(p.tgt_x,p.tgt_y,p.tgt_z,p.name) for p in valid]
        if not rows:QMessageBox.warning(self,"Nothing to export","There are no valid coordinates to export.");return
        path,_=QFileDialog.getSaveFileName(self,"Export Civil 3D CSV","Civil3D_Points.csv","CSV (*.csv)")
        if not path:return
        try:
            precision=current_precision(); ordered=order_points([PointResult(name,x,y,z,x,y,z,status="SUCCESS") for x,y,z,name in rows],mode=self._selected_order_mode())
            with open(path,"w",newline="",encoding="utf-8-sig") as f:
                writer=csv.writer(f);writer.writerow(["Point Number","Easting","Northing","Elevation","Description"])
                for item in ordered:
                    p=item.point; x=p.tgt_x if p.tgt_x is not None else p.src_x; y=p.tgt_y if p.tgt_y is not None else p.src_y; z=p.tgt_z if p.tgt_z is not None else p.src_z; writer.writerow([item.number,f"{x:.{precision}f}",f"{y:.{precision}f}",f"{(z or 0):.{precision}f}",p.name])
            QMessageBox.information(self,"Civil 3D CSV Exported",f"Civil 3D point file created:\n{path}")
        except Exception as exc:QMessageBox.critical(self,"CSV Export Error",str(exc))
