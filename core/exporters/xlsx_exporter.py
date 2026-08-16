"""XLSX exporter with coordinate-based sequential ordering."""
from __future__ import annotations
from datetime import date
from typing import List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from core.models import PointResult
from core.point_ordering import order_points

NAVY = "1F3864"; WHITE = "FFFFFF"; LGRAY = "F2F2F2"; GREEN = "C6EFCE"; RED = "FFC7CE"
thin = Side(style="thin", color="BFBFBF"); BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
APP_VERSION = "1.1.0"


def export_xlsx(points: List[PointResult], out_path: str, source_crs: str, target_crs: str,
                source_crs_details: Optional[dict] = None, precision: int = 3,
                order_mode: str = "GRID_ZIGZAG", tolerance: float | None = None,
                reverse: bool = False) -> str:
    precision = max(0, min(6, int(precision)))
    ordered = order_points(points, mode=order_mode, tolerance=tolerance, reverse=reverse)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Points"
    headers = ["No.", "Point Name", "Source X", "Source Y", "Source Z", "Target X", "Target Y", "Target Z", "Target X,Y", "Status"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h); c.font = Font(bold=True, color=WHITE); c.fill = PatternFill("solid", fgColor=NAVY); c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BORDER
    fmt = f"#,##0.{'0'*precision}" if precision > 0 else "#,##0"
    for item in ordered:
        p = item.point; row = [item.number, p.name, p.src_x, p.src_y, p.src_z, p.tgt_x, p.tgt_y, p.tgt_z,
                                f"{p.tgt_x:.{precision}f}, {p.tgt_y:.{precision}f}" if p.tgt_x is not None and p.tgt_y is not None else "", p.status]
        for j, v in enumerate(row, start=1):
            cell = ws.cell(row=item.number + 1, column=j, value=v); cell.border = BORDER
            if j in (3,4,5,6,7,8) and isinstance(v,(int,float)): cell.number_format = fmt
            if j == 10:
                if v == "SUCCESS": cell.fill = PatternFill("solid", fgColor=GREEN)
                elif v == "FAILED": cell.fill = PatternFill("solid", fgColor=RED)
    for col, w in zip("ABCDEFGHIJ", [6,20,14,14,10,14,14,10,20,10]): ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    info = wb.create_sheet("Project Info")
    rows = [("Source CRS", source_crs), ("Target CRS", target_crs), ("Transformation", "PROJ (pyproj.Transformer)"),
            ("Datum", (source_crs_details or {}).get("datum", "N/A")), ("Projection", (source_crs_details or {}).get("projection", "N/A")),
            ("Ellipsoid", (source_crs_details or {}).get("ellipsoid", "N/A")), ("Units", (source_crs_details or {}).get("units", "N/A")),
            ("Number of Points", len(points)), ("Ordering", order_mode), ("Date", date.today().isoformat()), ("Decimal Precision", precision),
            ("Software Version", f"MH GeoSuite Pro v{APP_VERSION}")]
    for i, (label, val) in enumerate(rows, start=1):
        lc=info.cell(row=i,column=1,value=label); lc.font=Font(bold=True); lc.fill=PatternFill("solid",fgColor=LGRAY); lc.border=BORDER
        vc=info.cell(row=i,column=2,value=val); vc.border=BORDER
    info.column_dimensions["A"].width=22; info.column_dimensions["B"].width=40
    wb.save(out_path); return out_path
