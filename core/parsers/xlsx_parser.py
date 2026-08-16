"""XLSX parser with automatic coordinate-column detection."""
from __future__ import annotations
from dataclasses import dataclass
from typing import List, Optional
import openpyxl
from core.models import PointResult

@dataclass
class ColumnMapping:
    name_col: Optional[str]
    x_col: str
    y_col: str
    z_col: Optional[str] = None

def _norm(s: object)->str:return "".join(ch for ch in str(s).strip().lower() if ch.isalnum())
def _find_col(headers:list[str], aliases:tuple[str,...], fallback:int|None=None)->int|None:
    ns=[_norm(h) for h in headers]; al=[_norm(a) for a in aliases]
    for a in al:
        if a in ns:return ns.index(a)
    for i,h in enumerate(ns):
        if any(a in h for a in al):return i
    return fallback
def sniff_columns(path: str, sheet: Optional[str] = None) -> List[str]:
    wb=openpyxl.load_workbook(path,read_only=True,data_only=True); ws=wb[sheet] if sheet else wb.active
    header=[str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1,max_row=1))]; wb.close()
    return header or ["Column 1","Column 2","Column 3"]
def parse_xlsx(path: str, mapping: ColumnMapping, sheet: Optional[str] = None) -> List[PointResult]:
    wb=openpyxl.load_workbook(path,read_only=True,data_only=True); ws=wb[sheet] if sheet else wb.active
    rows=list(ws.iter_rows(values_only=True)); wb.close()
    if not rows:return []
    headers=[str(h).strip() if h is not None else "" for h in rows[0]]; data=rows[1:]
    idx={h:i for i,h in enumerate(headers)}; points=[]
    if mapping.x_col not in idx or mapping.y_col not in idx:return parse_xlsx_auto(path,sheet)
    for i,row in enumerate(data,1):
        name=None
        if mapping.name_col and mapping.name_col in idx and idx[mapping.name_col]<len(row):name=str(row[idx[mapping.name_col]]).strip() if row[idx[mapping.name_col]] is not None else None
        try:x=float(row[idx[mapping.x_col]]);y=float(row[idx[mapping.y_col]])
        except (KeyError,TypeError,ValueError,IndexError):points.append(PointResult(name or f"PT-{i}",None,None,None,status="FAILED",message="Invalid or missing X/Y"));continue
        z=None
        if mapping.z_col and mapping.z_col in idx:
            try:z=float(row[idx[mapping.z_col]]) if row[idx[mapping.z_col]] is not None else None
            except (ValueError,TypeError):z=None
        points.append(PointResult(name or f"PT-{i}",x,y,z))
    return points
def parse_xlsx_auto(path: str, sheet: Optional[str] = None) -> List[PointResult]:
    wb=openpyxl.load_workbook(path,read_only=True,data_only=True); ws=wb[sheet] if sheet else wb.active
    rows=list(ws.iter_rows(values_only=True)); wb.close()
    if not rows:return []
    headers=[str(h).strip() if h is not None else "" for h in rows[0]]; data=rows[1:]
    xidx=_find_col(headers,("easting","east","x","xcoord","xcoordinate","longitude","lon"),None); yidx=_find_col(headers,("northing","north","y","ycoord","ycoordinate","latitude","lat"),None); zidx=_find_col(headers,("elevation","elev","height","z","zcoord","zcoordinate"),None); nidx=_find_col(headers,("pointnumber","pointno","pointid","pointcode","code","point","name","id","number"),None)
    if xidx is None or yidx is None:
        width=max((len(r) for r in data),default=0); numeric=[]
        for c in range(width):
            vals=[r[c] for r in data[:30] if len(r)>c and r[c] is not None]
            if vals:
                good=sum(isinstance(v,(int,float)) or (isinstance(v,str) and _is_number(v)) for v in vals)
                if good>=max(1,int(len(vals)*0.8)):numeric.append(c)
        if len(numeric)>=2:xidx,yidx=numeric[:2];zidx=numeric[2] if len(numeric)>=3 else None
    points=[]
    for i,row in enumerate(data,1):
        try:
            if xidx is None or yidx is None or len(row)<=max(xidx,yidx):raise ValueError
            x=float(row[xidx]);y=float(row[yidx]);z=float(row[zidx]) if zidx is not None and len(row)>zidx and row[zidx] is not None and str(row[zidx]).strip() else None
        except (ValueError,TypeError,IndexError):points.append(PointResult(f"PT-{i}",None,None,None,status="FAILED",message="Invalid or missing X/Y"));continue
        name=str(row[nidx]).strip() if nidx is not None and len(row)>nidx and row[nidx] is not None and str(row[nidx]).strip() else f"PT-{i}"
        points.append(PointResult(name,x,y,z))
    return points
def _is_number(v):
    try:float(v);return True
    except (ValueError,TypeError):return False
