#!/usr/bin/env python3
"""
Local smoke-test runner.

pytest cannot be installed in this offline sandbox (no network access to
PyPI). The tests/ directory contains the real pytest suite that WILL run
on the GitHub Actions Windows runner (which has internet access to
install pytest + all dependencies). This script re-exercises the same
core logic using only the Python standard library (unittest + tempfile)
so that the non-GUI, non-pyproj, non-ezdxf modules are genuinely verified
in this environment right now, rather than only claimed to work.

Modules requiring pyproj (CRS engine) or ezdxf (DXF export) or PySide6
(UI) are NOT exercised here — they are unavailable offline. They are
exercised by tests/test_crs_engine.py and tests/test_dxf_exporter.py on
CI, and the UI is exercised by the CI smoke test.
"""
import csv
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from core.models import PointResult
from core.parsers import csv_parser, xlsx_parser, kml_parser
from core.exporters import xlsx_exporter, csv_exporter
from core.validation.validator import validate_points, validate_zone_consistency
from core.batch.batch_processor import find_batch_files, run_batch, FileResult

import openpyxl


class TestCSVParser(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.path = Path(self.tmp.name) / "points.csv"
        with open(self.path, "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(["Point Name", "Easting", "Northing", "Elevation"])
            w.writerow(["P1", "46.84550218", "24.77373641", "650.5"])
            w.writerow(["P2", "", "24.79", ""])

    def tearDown(self):
        self.tmp.cleanup()

    def test_sniff_columns(self):
        self.assertEqual(
            csv_parser.sniff_columns(str(self.path)),
            ["Point Name", "Easting", "Northing", "Elevation"],
        )

    def test_parse_valid_row(self):
        mapping = csv_parser.ColumnMapping("Point Name", "Easting", "Northing", "Elevation")
        pts = csv_parser.parse_csv(str(self.path), mapping)
        self.assertAlmostEqual(pts[0].src_x, 46.84550218)
        self.assertEqual(pts[0].status, "PENDING")

    def test_bad_row_does_not_abort_batch(self):
        mapping = csv_parser.ColumnMapping("Point Name", "Easting", "Northing", "Elevation")
        pts = csv_parser.parse_csv(str(self.path), mapping)
        self.assertEqual(len(pts), 2)
        self.assertEqual(pts[1].status, "FAILED")


class TestXLSXParserAndExporter(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.path = Path(self.tmp.name) / "points.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Point Name", "Easting", "Northing", "Elevation"])
        ws.append(["P1", 46.84550218, 24.77373641, 650.5])
        wb.save(self.path)

    def tearDown(self):
        self.tmp.cleanup()

    def test_parse_xlsx(self):
        mapping = xlsx_parser.ColumnMapping("Point Name", "Easting", "Northing", "Elevation")
        pts = xlsx_parser.parse_xlsx(str(self.path), mapping)
        self.assertEqual(len(pts), 1)
        self.assertAlmostEqual(pts[0].src_x, 46.84550218)

    def test_export_xlsx_sheet_structure(self):
        out = Path(self.tmp.name) / "export.xlsx"
        pts = [PointResult("P1", 46.84, 24.77, 650.5, 500000.1, 2740123.4, 650.5, "SUCCESS")]
        xlsx_exporter.export_xlsx(pts, str(out), "EPSG:4326", "EPSG:20438", {"datum": "Ain el Abd 1970"})
        wb = openpyxl.load_workbook(out)
        self.assertEqual(wb.sheetnames, ["Points", "Project Info"])
        header = [c.value for c in next(wb["Points"].iter_rows(min_row=1, max_row=1))]
        self.assertEqual(
            header,
            ["No.", "Point Name", "Source X", "Source Y", "Source Z",
             "Target X", "Target Y", "Target Z", "Target X,Y", "Status"],
        )

    def test_export_csv(self):
        out = Path(self.tmp.name) / "export.csv"
        pts = [PointResult("P1", 46.84, 24.77, 650.5, 500000.1, 2740123.4, 650.5, "SUCCESS")]
        csv_exporter.export_csv(pts, str(out))
        content = out.read_text(encoding="utf-8-sig")
        self.assertIn("P1", content)
        self.assertIn("SUCCESS", content)


class TestKMLParser(unittest.TestCase):
    SAMPLE = b"""<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2"><Document>
<Placemark><name>Riyadh Tower</name>
<Point><coordinates>46.84550218,24.77373641,650.5</coordinates></Point>
</Placemark>
<Placemark><name>Corner A</name>
<MultiGeometry><Point><coordinates>46.85,24.78,0</coordinates></Point></MultiGeometry>
</Placemark>
</Document></kml>"""

    def test_parse_kml_bytes(self):
        points = kml_parser.parse_kml_bytes(self.SAMPLE)
        self.assertEqual(len(points), 2)
        self.assertEqual(points[0].name, "Riyadh Tower")
        self.assertAlmostEqual(points[0].src_x, 46.84550218)

    def test_kmz_auto_extraction(self):
        with tempfile.TemporaryDirectory() as d:
            kmz_path = Path(d) / "test.kmz"
            with zipfile.ZipFile(kmz_path, "w") as z:
                z.writestr("doc.kml", self.SAMPLE)
            points = kml_parser.parse_kmz_file(str(kmz_path))
            self.assertEqual(len(points), 2)

    def test_kmz_without_kml_raises(self):
        with tempfile.TemporaryDirectory() as d:
            kmz_path = Path(d) / "bad.kmz"
            with zipfile.ZipFile(kmz_path, "w") as z:
                z.writestr("readme.txt", "nope")
            with self.assertRaises(ValueError):
                kml_parser.parse_kmz_file(str(kmz_path))


class TestValidation(unittest.TestCase):
    def test_missing_coords(self):
        report = validate_points([PointResult("P1", None, None, None)])
        self.assertEqual(len(report.errors), 1)

    def test_out_of_range(self):
        report = validate_points([PointResult("P1", 200.0, 24.0, None)])
        self.assertTrue(any("Longitude" in e.message for e in report.errors))

    def test_duplicates_are_warnings_not_errors(self):
        report = validate_points([
            PointResult("P1", 1.0, 1.0, None),
            PointResult("P1", 1.0, 1.0, None),
        ])
        self.assertEqual(len(report.errors), 0)
        self.assertGreaterEqual(len(report.warnings), 1)

    def test_zone_mismatch_warning(self):
        w = validate_zone_consistency("UTM Zone 37N", "UTM Zone 38N")
        self.assertEqual(len(w), 1)


class TestBatchProcessor(unittest.TestCase):
    def test_batch_continues_after_failure(self):
        with tempfile.TemporaryDirectory() as d:
            for name in ["a.csv", "b.xlsx", "c.kmz", "d.kml"]:
                (Path(d) / name).write_text("x")
            (Path(d) / "ignored.txt").write_text("x")

            def process(path):
                if path.name == "b.xlsx":
                    raise RuntimeError("boom")
                return FileResult(str(path), "SUCCESS", 1, 1)

            files = find_batch_files(d)
            self.assertEqual(len(files), 4)  # ignored.txt excluded

            report = run_batch(d, process)
            self.assertEqual(report.success_count, 3)
            self.assertEqual(report.failed_count, 1)


if __name__ == "__main__":
    runner = unittest.TextTestRunner(verbosity=2)
    suite = unittest.TestLoader().loadTestsFromModule(sys.modules[__name__])
    result = runner.run(suite)
    sys.exit(0 if result.wasSuccessful() else 1)
